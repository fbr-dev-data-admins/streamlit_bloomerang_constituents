"""Excel file writing utilities using openpyxl."""

import logging
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

TEXT_FORMAT_COLUMNS = {
    "Home ZIP",
    "Home Phone",
    "Account Number",
    "Primary Contact First Name",
    "Primary Contact Last Name"
}


def write_excel(df: pd.DataFrame, filepath: str) -> None:
    """
    Write a DataFrame to an Excel file with proper formatting.

    Args:
        df: DataFrame to write (will be modified in place with fillna)
        filepath: Path to write the Excel file
    """
    bytes_data = df_to_excel_bytes(df)
    with open(filepath, "wb") as f:
        f.write(bytes_data)
    logger.info(f"Wrote Excel file: {filepath}")


def df_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Convert a DataFrame to Excel file bytes.

    Args:
        df: DataFrame to convert

    Returns:
        Raw bytes of the Excel file
    """
    df = df.fillna("").astype(str)

    wb = Workbook()
    ws = wb.active

    headers = list(df.columns)
    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    ws.freeze_panes = "A2"

    text_format_col_indices = set()
    for col_idx, header in enumerate(headers, 1):
        if header in TEXT_FORMAT_COLUMNS:
            text_format_col_indices.add(col_idx)

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in text_format_col_indices:
                cell.number_format = "@"

    for col_idx, header in enumerate(headers, 1):
        if col_idx in text_format_col_indices:
            ws.cell(row=1, column=col_idx).number_format = "@"

    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
